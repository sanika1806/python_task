import openpyxl
from openpyxl import Workbook
import os

# Define the file path for the Excel file
excel_file = "user_data.xlsx"

# Function to add a new user
def add_user():
    # Get user details
    name = input("Enter Name: ")
    email = input("Enter Email: ")
    phone = input("Enter Phone Number: ")
    
    # Check if the Excel file exists; if not, create it
    if not os.path.exists(excel_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        # Create headers
        sheet.append(["Name", "Email", "Phone Number"])
        workbook.save(excel_file)
    
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    
    # Append user data to the sheet
    sheet.append([name, email, phone])
    workbook.save(excel_file)
    print("User added successfully!")

# Function to display all users
def display_users():
    # Check if the Excel file exists
    if not os.path.exists(excel_file):
        print("No users found. Please add a user first.")
        return
    
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    
    # Display users in a readable format
    print("\nStored Users:")
    for row in sheet.iter_rows(values_only=True):
        print(f"Name: {row[0]}, Email: {row[1]}, Phone Number: {row[2]}")
    print()

# Main program loop
while True:
    print("Make a choice to proceed :-")

  
    print("Press '1' to Add user.")
    print("Press '2' to Display users.")
    print("press '3' to Exit.")
    
    choice = input("Press any key to continue :")
    
    if choice == "1":
        add_user()
    elif choice == "2":
        display_users()
    elif choice == "3":
        print("Exiting program.")
        break
    else:
        print("Invalid choice. Please select again.")
